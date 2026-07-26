"""
app/services/excel_service.py
--------------------------------
Excel import & export capabilities for the Equipment master, built on
openpyxl. Exports produce a styled workbook (header formatting, column
widths, freeze panes) suitable for offline review by site engineers;
imports validate rows and report per-row errors back to the caller.
"""

import io
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models.equipment import Equipment, EquipmentCategory, EquipmentStatus
from app.models.hierarchy import System
from app.models.vendor import Vendor
from app.models.user import User
from app.utils.access import has_project_access
EXPORT_HEADERS = [
    "Equipment Code", "Equipment Name", "System", "Category", "Vendor",
    "Drawing Number", "PO Number", "Status", "Installation Progress (%)",
    "Commissioning Progress (%)", "Assigned Engineer", "Expected Date",
    "Actual Date", "Remarks",
]

IMPORT_HEADERS = [
    "Equipment Code", "Equipment Name", "System Code", "Category Code",
    "Vendor Code", "Drawing Number", "PO Number", "Status",
    "Installation Progress (%)", "Commissioning Progress (%)",
    "Assigned Engineer Username", "Expected Date", "Actual Date", "Remarks",
]


class ExcelService:

    # ------------------------------------------------------------------
    # EXPORT
    # ------------------------------------------------------------------
    @staticmethod
    def export_equipment(equipment_list) -> io.BytesIO:
        """Build a formatted .xlsx workbook of the given Equipment records."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipment Master"

        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

        for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, eq in enumerate(equipment_list, start=2):
            values = [
                eq.equipment_code,
                eq.equipment_name,
                eq.system.name if eq.system else "",
                eq.category.name if eq.category else "",
                eq.vendor.name if eq.vendor else "",
                eq.drawing_number or "",
                eq.po_number or "",
                eq.status,
                eq.installation_progress or 0,
                eq.commissioning_progress or 0,
                eq.assigned_engineer.full_name if eq.assigned_engineer else "",
                eq.expected_date.strftime("%d-%b-%Y") if eq.expected_date else "",
                eq.actual_date.strftime("%d-%b-%Y") if eq.actual_date else "",
                eq.remarks or "",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Reasonable column widths
        widths = [16, 28, 20, 16, 20, 16, 16, 18, 14, 16, 18, 14, 14, 30]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_import_template() -> io.BytesIO:
        """Provide a blank template workbook with the expected import headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Template"
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, header in enumerate(IMPORT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ------------------------------------------------------------------
    # IMPORT
    # ------------------------------------------------------------------
    @staticmethod
    def import_equipment(file_stream) -> dict:
        """
        Parse an uploaded workbook (matching IMPORT_HEADERS) and upsert
        Equipment records. Returns a summary dict with counts and a list of
        row-level error messages so the UI can display exactly what failed.
        """
        wb = load_workbook(file_stream, data_only=True)
        ws = wb.active

        created, updated, errors = 0, 0, []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None for v in row):
                continue
            try:
                (equipment_code, equipment_name, system_code, category_code, vendor_code,
                 drawing_number, po_number, status, install_pct, commission_pct,
                 engineer_username, expected_date, actual_date, remarks) = (list(row) + [None] * 14)[:14]

                if not equipment_code or not equipment_name or not system_code:
                    errors.append(f"Row {row_idx}: Equipment Code, Name and System Code are required.")
                    continue

                system = System.query.filter_by(system_code=system_code).first()

                if not system:
                    errors.append(f"Row {row_idx}: System code '{system_code}' not found.")
                    continue

                # RBAC check
                project_id = system.area.unit.project_id

                if not has_project_access(project_id):
                    errors.append(
                        f"Row {row_idx}: You are not authorized to import equipment into project '{system.area.unit.project.project_code}'."
                    )
                    continue

                category = EquipmentCategory.query.filter_by(code=category_code).first() if category_code else None
                vendor = Vendor.query.filter_by(vendor_code=vendor_code).first() if vendor_code else None
                engineer = User.query.filter_by(username=engineer_username).first() if engineer_username else None

                equipment = Equipment.query.filter_by(
                    system_id=system.id, equipment_code=equipment_code
                ).first()

                is_new = equipment is None
                if is_new:
                    equipment = Equipment(system_id=system.id, equipment_code=equipment_code)

                equipment.equipment_name = equipment_name
                equipment.category_id = category.id if category else None
                equipment.vendor_id = vendor.id if vendor else None
                equipment.assigned_engineer_id = engineer.id if engineer else None
                equipment.drawing_number = drawing_number
                equipment.po_number = po_number
                equipment.status = status if status in EquipmentStatus.ALL else EquipmentStatus.NOT_STARTED
                equipment.installation_progress = float(install_pct) if install_pct else 0
                equipment.commissioning_progress = float(commission_pct) if commission_pct else 0
                equipment.remarks = remarks

                equipment.expected_date = ExcelService._parse_date(expected_date)
                equipment.actual_date = ExcelService._parse_date(actual_date)

                db.session.add(equipment)
                if is_new:
                    created += 1
                else:
                    updated += 1
            except Exception as exc:  # noqa: BLE001 - report any row failure gracefully
                errors.append(f"Row {row_idx}: {exc}")

        db.session.commit()
        return {"created": created, "updated": updated, "errors": errors}

    @staticmethod
    def _parse_date(value):
        if value is None or value == "":
            return None
        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value
        for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(value), fmt).date()
            except ValueError:
                continue
        return None
